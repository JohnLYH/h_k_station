# -*- coding: utf-8 -*-
import base64
import datetime
import json
import random
import time
from datetime import datetime as dt
from dateutil.relativedelta import relativedelta
import os
import shutil
import xlwt
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.writer.excel import save_virtual_workbook
import zipfile

from odoo import http
from odoo.http import request
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.writer.excel import save_virtual_workbook

from .mobile_api import to_json

EXPORT_TITLE_LIST = [
    'Work Order Description', 'Standard Job Description', 'Planned Start Date', 'Planned Completion Date',
    'Appointed Completion Date', 'Actual Start Date', 'Actual Complete Date', 'Group', 'Executor'
]

OTHER_ROW_1_LIST = ['EQUIPMENT No.', 'EQUIPMENT', 'BRAND', 'MODEL', 'SERIAL_NO', 'MANUAL REF. NO.', 'EQUIPMENT OWNER',
                    'LOCATION OF EQUIPMENT', 'FREQ. OF CAL.', 'CALIBRATION BODY',
                    'CALIBRATION REQUIPEMNETS  (SEE NOTE)', 'LAST MAINTENANCE DATE',
                    'MAINTENANCE DUE DATE', 'STATUS', 'REMARK']

APP_DIR = os.path.dirname(os.path.dirname(__file__))


class MaintenancePlan(http.Controller):

    @staticmethod
    def get_row1_list_colnum(row1_list, value):
        '''
        獲取list中某詞的index并+1，若沒有則返回None
        :param row1_list:
        :param value:
        :return:
        '''
        low_value = value.replace(' ', '').lower()
        try:
            return {'col': value, 'num': row1_list.index(low_value) + 1}
        except ValueError as e:
            return {'col': value, 'num': None}

    @staticmethod
    def excel_validate(sheet, num, cols_list, style):
        '''
        驗證excel的一行的錯誤性
        :param sheet: sheet活動表
        :param num: 行數，1開始
        :param cols_list: 需要驗證的列數列表
        :param style: sheet顏色style
        :return:
        '''
        row_error = False
        for col in cols_list:
            if sheet.cell(num, col).value is None:
                sheet.cell(num, col).fill = style
                row_error = True
        return row_error

    @http.route('/maintenance_plan/put_in_excel', type='http', csrf=False, auth='user')
    def put_in_excel(self, **kwargs):
        '''
        維修計劃管理頁面導入excel按鈕
        :param kwargs: excel的file信息
        :return:
        '''
        file = kwargs['file']
        filename = kwargs['file'].filename
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        error = False
        n_row = 0
        for row in sheet.rows:
            n_row += 1
            if n_row == 1:
                title_list = [None if i.value is None else i.value.replace(' ', '').lower() for i in row]
                work_num = self.get_row1_list_colnum(title_list, 'Work Order Number')  # 工單編號WorkOrderNumber
                work_order_type = self.get_row1_list_colnum(title_list, 'Work Nature L1')  # 工單類型WorkNatureL1
                equipment_num = self.get_row1_list_colnum(title_list, 'Equipment Number')  # 設備編號EquipmentNumber
                standard_job = self.get_row1_list_colnum(title_list, 'Standard Job Code')  # 標準工作StandardJobCode
                standard_job_description = self.get_row1_list_colnum(title_list,
                                                                     'Standard Job Description')  # 標準工作描述StandardJobDescription
                work_order_description = self.get_row1_list_colnum(title_list,
                                                                   'Work Order Description')  # 工單描述WorkOrderDescription
                plan_start_time = self.get_row1_list_colnum(title_list, 'Planned Start   Date')  # 建議開始時間PlannedStartDate
                plan_end_time = self.get_row1_list_colnum(title_list,
                                                          'Planned Completed Date')  # 建議結束時間PlannedCompletedDate
                # 檢查是否有colnum未存在的列
                none_col_list = []
                col_list = [work_num, work_order_type, equipment_num, standard_job, standard_job_description,
                            work_order_description, plan_start_time, plan_end_time]
                for check_col in col_list:
                    if check_col['num'] is None:
                        none_col_list.append(check_col['col'])
                # 如果有未存在的列，返回error
                if len(none_col_list) > 0:
                    return to_json({'error': True, 'message': '{}列不存在'.format([col for col in none_col_list])})
                else:
                    col_index = [col['num'] for col in col_list]
                    # 单元格背景色紅色
                    red_style = PatternFill(fill_type='solid', fgColor="FF3030")
                    # 单元格背景色綠色
                    green_style = PatternFill(fill_type='solid', fgColor="458B74")
            else:
                # 校驗特定列是否有空值或錯值
                row_error = self.excel_validate(sheet, n_row, col_index, red_style)
                if row_error is False:
                    work_order_count = request.env['maintenance_plan.maintenance.plan'].search_count([
                        ('num', '=', sheet.cell(n_row, work_num['num']).value)
                    ])
                    equipment_record = request.env['maintenance_plan.equipment'].search([
                        ('num', '=', sheet.cell(n_row, equipment_num['num']).value)
                    ])
                    standard_job_record = request.env['maintenance_plan.standard.job'].search([
                        ('name', '=', standard_job['num'])
                    ])
                    # 檢查是否已經存在工單
                    if work_order_count != 0 and sheet.cell(n_row, 1).value is not None:
                        sheet.cell(n_row, work_num['num']).fill = green_style
                        error = True
                    # 檢查是否存在對應編號設備
                    if len(equipment_record) == 0:
                        sheet.cell(n_row, equipment_num['num']).fill = red_style
                        error = True
                    # 檢查標準工作是否已經存在
                    if len(standard_job_record) == 0 and work_order_count == 0 and len(equipment_record) != 0:
                        standard_job_record = request.env['maintenance_plan.standard.job'].create({
                            'name': sheet.cell(n_row, standard_job['num']).value,
                            'description': sheet.cell(n_row, standard_job_description['num']).value
                        })
                    if work_order_count == 0 and len(equipment_record) != 0:
                        request.env['maintenance_plan.maintenance.plan'].create({
                            'num': sheet.cell(n_row, work_num['num']).value,
                            'work_order_type': sheet.cell(n_row, work_order_type['num']).value,
                            'work_order_description': sheet.cell(n_row, work_order_description['num']).value,
                            'standard_job_id': standard_job_record.id,
                            'equipment_id': equipment_record.id,
                            'plan_start_time': sheet.cell(n_row, plan_start_time['num']).value.split(' ')[0],
                            'plan_end_time': sheet.cell(n_row, plan_end_time['num']).value.split(' ')[0],
                        })
                else:
                    error = True
        if error is True:
            new_file = request.env['maintenance_plan.trans.excel'].create({
                'name': filename.split('.')[0],
                'file': save_virtual_workbook(workbook)
            })
            return to_json({'error': error, 'message': '文件有部分錯誤信息，請修改后再次傳入', 'file_id': new_file.id})
        else:
            return to_json({'error': error, 'message': '上傳成功'})

    @http.route('/maintenance_plan/equipment_put_in_excel', type='http', csrf=False, auth='user')
    def equipment_put_in_excel(self, **kwargs):
        '''
        設備管理頁面上傳設備excel
        :param kwargs:
        :return:
        '''
        file = kwargs['file']
        filename = kwargs['file'].filename
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        error = False
        n_row = 0
        qr_code_record_ids = []
        for row in sheet.rows:
            n_row += 1
            if n_row == 1:
                title_list = [None if i.value is None else i.value.replace(' ', '').lower() for i in row]
                equipment_num = self.get_row1_list_colnum(title_list, 'Equipment No.')  # 設備編號
                parent_equipment_num = self.get_row1_list_colnum(title_list, 'parent_equipment_NO.')  # 父設備編號
                serial_num = self.get_row1_list_colnum(title_list, 'Serial No.')  # 序列號
                detailed_location = self.get_row1_list_colnum(title_list, 'LOCATION CODE')  # 詳細位置
                equipment_type = self.get_row1_list_colnum(title_list, 'Equipment Class')  #設備類別
                equipment_model = self.get_row1_list_colnum(title_list, 'Model')  # 設備型號
                description = self.get_row1_list_colnum(title_list, 'Equipment Description')  # 設備名稱(即設備描述)
                status = self.get_row1_list_colnum(title_list, 'status')  # 設備狀態
                item_code = self.get_row1_list_colnum(title_list, 'Item Code')  # 庫存編碼
                line = self.get_row1_list_colnum(title_list, 'Line')  # 線別
                station = self.get_row1_list_colnum(title_list, 'station')  # 站點
                direction = self.get_row1_list_colnum(title_list, 'Direction')  # 方向
                start_chainage = self.get_row1_list_colnum(title_list, 'Chainage-start')  # 起始公里標
                end_chainage = self.get_row1_list_colnum(title_list, 'Chainage-end')  # 終點公里標
                last_installation_date = self.get_row1_list_colnum(title_list, ' Last installation date')  # 最後安裝日期
                service_since = self.get_row1_list_colnum(title_list, 'service date/service since')  # 啟用時間
                expected_asset_life = self.get_row1_list_colnum(title_list, 'Expected asset life')  # 預計使用時間
                warranty = self.get_row1_list_colnum(title_list, 'Warranty')  # 質保期
                supplier = self.get_row1_list_colnum(title_list, 'Supplier')  # 供應商
                oem_manufacturer = self.get_row1_list_colnum(title_list, 'OEM Manufacturer')  # 原始設備製造商
                lead_maintainer = self.get_row1_list_colnum(title_list, ' lead maintainer')  # 設備維護者
                # 檢查是否有colnum未存在的列
                none_col_list = []
                col_list = [equipment_num, parent_equipment_num, serial_num, detailed_location, equipment_type,
                            equipment_model, description, status, item_code, line, station, direction, start_chainage,
                            end_chainage, last_installation_date, last_installation_date, service_since,
                            expected_asset_life, warranty, supplier, oem_manufacturer, lead_maintainer]
                for check_col in col_list:
                    if check_col['num'] is None:
                        none_col_list.append(check_col['col'])
                # 如果有未存在的列，返回error
                if len(none_col_list) > 0:
                    return to_json({'error': True, 'message': '{}列不存在'.format([col for col in none_col_list])})
                else:
                    col_index = [serial_num['num'], equipment_model['num'], equipment_type['num'], status['num'],
                                 item_code['num']]
                    # 单元格背景色紅色
                    red_style = PatternFill(fill_type='solid', fgColor="FF3030")
                    # 单元格背景色綠色
                    green_style = PatternFill(fill_type='solid', fgColor="458B74")
            else:
                # 校驗特定列是否有空值或錯值
                row_error = self.excel_validate(sheet, n_row, col_index, red_style)
                # 檢查狀態是否符合選擇項
                if sheet.cell(n_row, status['num']).value not in ['Expired', 'Effective']:
                    sheet.cell(n_row, status['num']).fill = red_style
                    error = True
                    row_error = True
                if row_error is False:
                    serial_number_record = request.env['maintenance_plan.equipment'].search([
                        ('serial_number', '=', sheet.cell(n_row, serial_num['num']).value)
                    ])
                    equipment_type_record = request.env['maintenance_plan.equipment.type'].search([
                        ('name', '=', sheet.cell(n_row, equipment_type['num']).value)
                    ])
                    equipment_model_record = request.env['maintenance_plan.equipment_model'].search([
                        ('equipment_model', '=', sheet.cell(n_row, equipment_model['num']).value)
                    ])
                    # 檢查設備類型是否存在，不存在則標記，不允許創建
                    if len(equipment_type_record) == 0:
                        sheet.cell(n_row, equipment_type['num']).fill = red_style
                        error = True
                    # 檢查序列號是否存在，存在則跳過并標綠
                    if len(serial_number_record) != 0:
                        sheet.cell(n_row, serial_num['num']).fill = green_style
                        error = True
                        continue
                    # 檢查設備型號是否存在，不存在則創建
                    if len(equipment_model_record) == 0 and len(equipment_type_record) != 0:
                        equipment_model_record = request.env['maintenance_plan.equipment_model'].create({
                            'equipment_model': sheet.cell(n_row, equipment_model['num']).value,
                            'description': sheet.cell(n_row, description['num']).value
                        })
                    # 創建設備記錄
                    if len(equipment_type_record) != 0:
                        eq = request.env['maintenance_plan.equipment'].create({
                            'num': sheet.cell(n_row, equipment_num['num']).value,
                            'parent_equipment_num': sheet.cell(n_row, parent_equipment_num['num']).value,
                            'serial_number_id': request.env['maintenance_plan.equipment.serial_number'].create({
                                'num': sheet.cell(n_row, serial_num['num']).value
                            }).id,
                            'equipment_type_id': equipment_type_record.id,
                            'line': sheet.cell(n_row, line['num']).value,
                            'station': sheet.cell(n_row, station['num']).value,
                            'equipment_model': equipment_model_record.id,
                            'status': sheet.cell(n_row, status['num']).value,
                            'item_code': sheet.cell(n_row, item_code['num']).value,
                            'direction': sheet.cell(n_row, direction['num']).value,
                            'start_chainage': sheet.cell(n_row, start_chainage['num']).value,
                            'end_chainage': sheet.cell(n_row, end_chainage['num']).value,
                            'detailed_location': sheet.cell(n_row, detailed_location['num']).value,
                            'last_installation_date': sheet.cell(n_row, last_installation_date['num']).value,
                            'service_since': sheet.cell(n_row, service_since['num']).value,
                            'expected_asset_life': sheet.cell(n_row, expected_asset_life['num']).value,
                            'warranty': sheet.cell(n_row, warranty['num']).value,
                            'supplier': sheet.cell(n_row, supplier['num']).value,
                            'oem_manufacturer': sheet.cell(n_row, oem_manufacturer['num']).value,
                            'lead_maintainer': sheet.cell(n_row, lead_maintainer['num']).value,
                        })
                        qr_code_record_ids.append(eq.id)
                else:
                    error = True
        if error is True:
            new_file = request.env['maintenance_plan.trans.excel'].create({
                'name': filename.split('.')[0],
                'file': save_virtual_workbook(workbook)
            })
            return to_json({'error': error, 'message': '文件有部分錯誤信息，請修改后再次傳入', 'file_id': new_file.id,
                            'qr_code_record_ids': qr_code_record_ids})
        else:
            return to_json({'error': error, 'message': '上傳成功', 'qr_code_record_ids': qr_code_record_ids})

    @http.route('/maintenance_plan/down_wrong_file', type='http', auth="user", methods=['GET'])
    def down_wrong_file(self, **kwargs):
        '''
        返回错误excel内容
        :param kwargs: file_id
        :return:
        '''
        file_id = int(kwargs['file_id'])
        wb = request.env['maintenance_plan.trans.excel'].browse(file_id)
        response = request.make_response(wb.file)
        response.headers["Content-Disposition"] = "attachment; filename={}". \
            format((wb.name + '錯誤.xlsx').encode().decode('latin-1'))
        return response

    @http.route('/maintenance_plan/export_work_order', auth='user', csrf=False, type='http', method=['POST'])
    def export_work_order(self, **kwargs):
        '''
        工單管理頁面導出excel
        :param kwargs:
        :return:
        '''
        domains = json.loads(kwargs['domain'])
        limit = kwargs['limit']
        offset = kwargs['offset']
        records = request.env['maintenance_plan.maintenance.plan'].search(
            [i[0] for i in domains], limit=int(limit), offset=int(offset) - 1
        )
        wb = openpyxl.Workbook()
        sheet = wb.active
        for num in range(0, len(EXPORT_TITLE_LIST)):
            sheet.cell(1, num + 1).value = EXPORT_TITLE_LIST[num]
        row_num = 1
        for record in records:
            row_num += 1
            sheet.cell(row_num, 1).value = record.work_order_description  # 工单描述
            sheet.cell(row_num, 2).value = record.standard_job_id.name  # 标准工作
            sheet.cell(row_num, 3).value = record.plan_start_time  # 建议时间(开始)
            sheet.cell(row_num, 4).value = record.plan_end_time  # 建议时间(开始)
            sheet.cell(row_num, 5).value = record.action_time or None  # 計劃執行時間
            sheet.cell(row_num, 6).value = record.actual_start_time or None  # 实际开始时间
            sheet.cell(row_num, 7).value = record.actual_end_time or None  # 实际结束时间
            sheet.cell(row_num, 8).value = record.action_dep_id.name or None  # 执行班组
            sheet.cell(row_num, 9).value = record.executor_id.name or None  # 执行人
        response = request.make_response(save_virtual_workbook(wb))
        return response

    @http.route('/maintenance_plan/export_qr_code_zip', auth='user', csrf=False, type='http', method=['POST'])
    def export_qr_code_zip(self, **kwargs):
        qr_list = json.loads(kwargs['qr_list'])
        now_day = dt.strftime(dt.now() + relativedelta(hours=8), '%Y-%m-%d')
        zip_name = '{}{}/new.zip'.format(now_day, '导入设备二维码')
        if qr_list is not None:
            equipment_records = request.env['maintenance_plan.equipment'].browse(qr_list)
            file_name = os.path.join(APP_DIR, 'static/trans_zip', str(int(time.time())) + str(random.randint(1, 1000)))
            # 創建臨時文件夾
            os.mkdir(file_name)
            # 創建臨時壓縮文件
            new_zip = zipfile.ZipFile(os.path.join(file_name, 'new.zip'), 'w')
            # 打包二維碼圖片
            for record in equipment_records:
                # 圖片保存路徑
                image_path = os.path.join(file_name, '{}+{}.png'.format(record.num or 'none', record.serial_number))
                with open(image_path, "wb") as f:
                    # 加入壓縮文件，并更換在壓縮文件中的路徑及名稱
                    f.write(base64.b64decode(record.serial_number_id.qr_code))
                new_zip.write(image_path, '{}/{}+{}.png'.format(zip_name, record.num or 'none', record.serial_number),
                              compress_type=zipfile.ZIP_DEFLATED)
            new_zip.close()
            # 讀取壓縮文件內容
            with open(os.path.join(file_name, 'new.zip'), 'rb') as zip_f:
                response = request.make_response(zip_f.read())
            # 刪除臨時文件夾
            shutil.rmtree(file_name)
            return response
        # TODO: 導出設備二維碼

    @http.route('/maintenance_plan/materials_upload_files', auth='user', csrf=False, methods=['POST'])
    def materials_upload_files(self, **kw):
        try:
            # 設備類型
            field_type = kw['field_type']
            edition = kw['edition']
            file = kw['file']
            # 編號
            numbering = kw['numbering']
            select_file_name = file.filename
            # 設備型號
            equipment_id = kw['id']
            # 上傳的文件時pdf還是視頻
            upload_tpye = kw['upload_tpye']
            # 判斷這個型號下是否有相同的版本號或者編號存在
            # maintenance_plan.reference_materials_manage
            materials_manage_count = request.env['maintenance_plan.reference_materials_manage'].search_count([
                ('equipment_id', '=', equipment_id), ('field_type', '=', field_type), '|',
                ('numbering', '=', numbering), ('edition', '=', edition)])
            if materials_manage_count > 0:
                return json.dumps({'error': 1, 'message': '上傳失敗,已存在相同的編號或者版本'})
            file_name = str(int(time.time())) + str(random.randint(1, 1000)) + select_file_name
            file.save(file_name)
            with open(file_name, "rb") as e:
                b64str = base64.b64encode(e.read())
            os.remove(file_name)
            values = {
                'equipment_id': int(equipment_id),
                'field_type': field_type,
                'edition': edition,
                'numbering': numbering,
                'select_file': b64str,
                'select_file_name': select_file_name,
                'upload_tpye': upload_tpye
            }
            equipment_model = request.env['maintenance_plan.equipment_model'].sudo().search([('id', '=', equipment_id)])
            if upload_tpye == 'mp4':
                values['status'] = True
                equipment_model.write({'reference_materials_manage_ids': [(0, 0, values)]})
            else:
                equipment_model.write({'reference_materials_manage_ids': [(0, 0, values)]})
                # TODO：生成審批記錄
            try:
                # 變更原因
                reasons_change = kw['reasons_change']
            except:
                reasons_change = ''
            try:
                # 變更細節
                reasons_details = kw['reasons_details']
            except:
                reasons_details = ''
            operation_type = '新增'
            user_id = request.uid
            operation_time = datetime.datetime.now()
            values = {
                'reasons_change': reasons_change,
                'reasons_details': reasons_details,
                'operation_type': operation_type,
                'user_id': user_id,
                'operation_time': operation_time,
                'field_type': field_type,
                'select_file_name': select_file_name,
                'edition': edition,
                'numbering': numbering,
            }
            equipment_model.write({'reference_materials_manage_records': [(0, 0, values)]})
        except:
            return json.dumps({'error': 1, 'message': '上傳失敗'})
        return json.dumps({'error': 0})

    @http.route('/maintenance_plan/materials_change', auth='user', csrf=False, methods=['POST'])
    def materials_change(self, **kw):
        try:
            # 設備類型
            field_type = kw['field_type']
            edition = kw['edition']
            # 編號
            numbering = kw['numbering']
            # 設備型號
            equipment_id = kw['res_id']
            # 上傳的文件是視頻還是pdf(pdf需要生成審批記錄,視頻不需要)
            upload_tpye = kw['upload_tpye']
            reference_materials_manage_id = kw['id']
            # 判斷這個型號下是否有相同的版本號或者編號存在
            materials_manage_count = request.env['maintenance_plan.reference_materials_manage'].search_count([
                ('equipment_id', '=', equipment_id), ('field_type', '=', field_type),
                ('id', '!=', reference_materials_manage_id), '|',
                ('numbering', '=', numbering), ('edition', '=', edition)])
            if materials_manage_count > 0:
                return json.dumps({'error': 1, 'message': '上傳失敗,已存在相同的編號或者版本'})
            if kw['file'] == 'undefined':
                values = {
                    'field_type': field_type,
                    'edition': edition,
                    'numbering': numbering,
                    'upload_tpye': upload_tpye
                }
                select_file_name = request.env['maintenance_plan.reference_materials_manage'].sudo().search(
                    [('id', '=', reference_materials_manage_id)]).select_file_name
            else:
                file = kw['file']
                select_file_name = file.filename
                file_name = str(int(time.time())) + str(random.randint(1, 1000)) + select_file_name
                file.save(file_name)
                with open(file_name, "rb") as e:
                    b64str = base64.b64encode(e.read())
                os.remove(file_name)
                values = {
                    'equipment_id': int(equipment_id),
                    'field_type': field_type,
                    'edition': edition,
                    'numbering': numbering,
                    'select_file': b64str,
                    'select_file_name': select_file_name,
                    'upload_tpye': upload_tpye
                }
            equipment_model = request.env['maintenance_plan.equipment_model'].sudo().search([('id', '=', equipment_id)])
            if upload_tpye == 'mp4':
                values['status'] = True
            equipment_model.write({'reference_materials_manage_ids': [(1, reference_materials_manage_id, values)]})
            # TODO：生成審批記錄
            # 生成記錄
            try:
                # 變更原因
                reasons_change = kw['reasons_change']
            except:
                reasons_change = ''
            try:
                # 變更細節
                reasons_details = kw['reasons_details']
            except:
                reasons_details = ''
            operation_type = '修改'
            user_id = request.uid
            operation_time = datetime.datetime.now()
            values = {
                'reasons_change': reasons_change,
                'reasons_details': reasons_details,
                'operation_type': operation_type,
                'user_id': user_id,
                'operation_time': operation_time,
                'field_type': field_type,
                'select_file_name': select_file_name,
                'edition': edition,
                'numbering': numbering,
            }
            equipment_model.write({'reference_materials_manage_records': [(0, 0, values)]})
        except:
            return json.dumps({'error': 1, 'message': '上傳失敗'})
        return json.dumps({'error': 0})


class OtherEquipment(http.Controller):
    @staticmethod
    def excel_validate(request, new_sheet, num, cols_list, style):
        '''
        驗證excel的一行的錯誤性
        :param sheet: sheet活動表
        :param num: 行數，1開始
        :param cols_list: 需要驗證的列數列表
        :param style: sheet顏色style
        :param has_date_col: 是否驗證第17、18列的時間格式
        :return:
        '''
        row_error = False
        for col in cols_list:

            if new_sheet.cell(num, col).value is None:
                pass
            elif col == 7:
                departments = new_sheet.cell(num, col).value
                if not request.env['user.department'].sudo().search([('name', '=', departments)]):
                    new_sheet.cell(num, col).fill = style
                    row_error = True
                else:
                    pass
            else:
                freq_of_cal = new_sheet.cell(num, col).value
                if freq_of_cal == 'ON\nCONDITION':
                    pass
                else:
                    try:
                        freq_of_cal = int(freq_of_cal)
                    except:
                        new_sheet.cell(num, col).fill = style
                        row_error = True
        return row_error

    @staticmethod
    def check_date(new_sheet, num, cols_list, style):
        row_error = False
        for col in cols_list:
            if new_sheet.cell(num, col).value is None:
                pass
            else:
                print(new_sheet.cell(num, col).value)
                try:
                    dt.strptime(new_sheet.cell(num, col).value, '%Y/%m/%d')
                except Exception as e:
                    try:
                        dt.strptime(new_sheet.cell(num, col).value, '%Y-%m-%d')
                    except:
                        new_sheet.cell(num, col).fill = style
                        row_error = True
        return row_error

    @staticmethod
    def check_none(new_sheet, num, cols_list, style):
        row_error = False
        for col in cols_list:
            if new_sheet.cell(num, col).value is None:
                new_sheet.cell(num, col).fill = style
                row_error = True
            if col == 1:
                equipment_num = new_sheet.cell(num, col).value
                if new_sheet.cell(num, 7).value:
                    departments = new_sheet.cell(num, 7).value
                    # 檢查當前組是否有這個設備編號了
                    if request.env['maintenance_plan.other_equipment'].sudo().search_count(
                            [('equipment_num', '=', equipment_num),
                             ('departments.name', '=', departments)]) > 0:
                        new_sheet.cell(num, col).fill = style
                        row_error = True
                else:
                    if request.env['maintenance_plan.other_equipment'].sudo().search_count(
                            [('equipment_num', '=', equipment_num)]) > 0:
                        new_sheet.cell(num, col).fill = style
                        row_error = True
            if col == 14:
                status = new_sheet.cell(num, col).value
                if status not in ['OK', 'Expired', 'Scraped']:
                    new_sheet.cell(num, col).fill = style
                    row_error = True
        return row_error

    @staticmethod
    def set_excel_style_color(colour_map_key):
        '''
        設置單元格背景色
        :param colour_map_key: 顏色對應鍵值對的key
        :return:
        '''
        style = xlwt.XFStyle()
        pattern = xlwt.Pattern()
        pattern.pattern = 1
        pattern.pattern_fore_colour = xlwt.Style.colour_map[colour_map_key]
        style.pattern = pattern
        return style

    @http.route('/other_equipment/put_in_excel', type='http', csrf=False, auth='user')
    def put_in_excel(self, **kwargs):
        '''
        工器具管理導入excel按鈕
        :param kwargs: excel的file信息
        :return:
        '''
        try:
            file = kwargs['file']
            filename = kwargs['file'].filename
            workbook = load_workbook(file, data_only=True)
            sheet = workbook.active
            # 设置单元格背景色紅色
            red_style = PatternFill(fill_type='solid', fgColor="FF3030")
            # 设置单元格背景色綠色
            green_style = PatternFill(fill_type='solid', fgColor="458B74")
            error = False
            n_row = 0
            for row in sheet.rows:
                n_row += 1
                if n_row == 1:
                    if [i.value for i in row] != OTHER_ROW_1_LIST:
                        return json.dumps({'message': '表格不符', 'error': True})
                else:
                    # 這裡是檢測日期格式是否正確
                    row_error = self.check_date(sheet, n_row, [12,13], red_style)
                    # 檢驗三個必填字段是否為空
                    row_error2 = self.check_none(sheet, n_row, [1, 2], red_style)
                    # 這裡檢查參數是否合格
                    row_error3 = self.excel_validate(request, sheet, n_row,
                                                     [7, 9, 14], red_style)
                    if row_error or row_error2 or row_error3:
                        error = True
                    else:
                        # 設備編號
                        equipment_num = sheet.cell(n_row, 1).value
                        # 設備名稱
                        equipment_name = sheet.cell(n_row, 2).value
                        # 品牌
                        brand = sheet.cell(n_row, 3).value
                        # 型號
                        model = sheet.cell(n_row, 4).value
                        # 序列號
                        serial_no = sheet.cell(n_row, 5).value
                        # 參考手冊編號
                        manual_ref_no = sheet.cell(n_row, 6).value
                        # 設備擁有者
                        equipment_owner = sheet.cell(n_row, 7).value
                        # 設備位置
                        location_of_equipment = sheet.cell(n_row, 8).value
                        # 檢驗週期
                        freq_of_cal = sheet.cell(n_row, 9).value
                        # 檢驗主體
                        calibration_body = sheet.cell(n_row, 10).value
                        # 檢驗要求
                        calibration_requipemnets = sheet.cell(n_row, 11).value
                        # 最後維護日期
                        last_maintenance_date = sheet.cell(n_row, 12).value
                        # 應用到期時間
                        try:
                            mymonth = int(freq_of_cal)
                        except:
                            mymonth = False
                        if mymonth:
                            maintenance_due_data = dt.strptime(last_maintenance_date, '%Y/%m/%d') + relativedelta(
                                months=mymonth)
                        else:
                            maintenance_due_data = sheet.cell(n_row, 13).value
                        # 狀態
                        status = sheet.cell(n_row, 14).value
                        # 備註
                        remark = sheet.cell(n_row, 15).value
                        equipment_owner = request.env['user.department'].sudo().search(
                            [('name', '=', equipment_owner)]).id
                        records = request.env['maintenance_plan.other_equipment'].create({
                            'departments': equipment_owner, 'equipment_num': equipment_num,
                            'equipment_name': equipment_name, 'brand': brand, 'model': model,
                            'serial_no': serial_no, 'manual_ref_no': manual_ref_no,
                            'location_of_equipment': location_of_equipment,
                            'freq_of_cal': freq_of_cal, 'calibration_body': calibration_body,
                            'calibration_requipemnets': calibration_requipemnets,
                            'last_maintenance_date': last_maintenance_date, 'status': status,
                            'maintenance_due_data': maintenance_due_data, 'type': 'create',
                        })
                        request.env['maintenance_plan.other_equipment_records'].create({
                            'other_equipment_id': records.id,
                            'operation_time': datetime.datetime.now(),
                            'operation_type': '新建',
                            'content': '工器具新增',
                            'user_id': request.uid,
                            'remark': remark,
                        })
                        error = False
                    if row_error is True:
                        error = True
            if error is True:
                new_file = request.env['maintenance_plan.trans.excel'].create({
                    'name': filename.split('.')[0],
                    'file': save_virtual_workbook(workbook)
                })
                return json.dumps({'error': error, 'message': '文件有部分錯誤信息，請修改后再次傳入', 'file_id': new_file.id})
            else:
                return json.dumps({'error': error, 'message': '上傳成功'})
        except:
            error = True
            return json.dumps({'error': error, 'message': '上傳失敗'})

    @http.route('/other_equipment/get_in_excel', type='http', website=True, csrf=False, auth='user')
    def get_in_excel(self, domains):
        '''
        工器具管理導出excel按鈕
        :param kwargs: 需要導出的domian條件
        :return:
        '''
        arr = []
        # try:
        if domains:
            x = domains.split(',')
            num = int(len(x) / 3 + 1)
            for i in range(1, num):
                tuples = (x[3 * i - 3], x[3 * i - 2], x[3 * i - 1])
                arr.append(tuples)
            other_equipments = request.env['maintenance_plan.other_equipment'].search(arr)
        else:
            other_equipments = request.env['maintenance_plan.other_equipment'].search([])
        wb = Workbook()
        # 激活 worksheet
        ws = wb.active
        ws.append(OTHER_ROW_1_LIST)
        len_other_equipments = len(other_equipments)
        for i in range(1, len_other_equipments + 1):
            my_records = [
                other_equipments[i - 1].equipment_num if other_equipments[i - 1].equipment_num else '',
                other_equipments[i - 1].equipment_name if other_equipments[i - 1].equipment_name else '',
                other_equipments[i - 1].brand if other_equipments[i - 1].brand else '',
                other_equipments[i - 1].model if other_equipments[i - 1].model else '',
                other_equipments[i - 1].serial_no if other_equipments[i - 1].serial_no else '',
                other_equipments[i - 1].manual_ref_no if other_equipments[i - 1].manual_ref_no else '',
                other_equipments[i - 1].departments.name if other_equipments[i - 1].departments.name else '',
                other_equipments[i - 1].location_of_equipment if other_equipments[
                    i - 1].location_of_equipment else '',
                other_equipments[i - 1].freq_of_cal if other_equipments[i - 1].freq_of_cal else '',
                other_equipments[i - 1].calibration_body if other_equipments[i - 1].calibration_body else '',
                other_equipments[i - 1].calibration_requipemnets if other_equipments[
                    i - 1].calibration_requipemnets else '',
                other_equipments[i - 1].last_maintenance_date if other_equipments[
                    i - 1].last_maintenance_date else '',
                other_equipments[i - 1].maintenance_due_data if other_equipments[i - 1].maintenance_due_data else '',
                other_equipments[i - 1].status if other_equipments[i - 1].status else '',
                other_equipments[i - 1].remark if other_equipments[i - 1].remark else '',
            ]
            ws.append(my_records)
        print('xxx')
        file_name = str(time.time()) + str(random.randint(1, 10000)) + "sample.xlsx"
        wb.save(file_name)
        with open(file_name, 'rb') as f:
            data = f.read()
            new_file = request.env['maintenance_plan.trans.excel'].create({
                'name': '工器具導出數據',
                'file': data
            })
            os.remove(file_name)
        return json.dumps({'error': 0, 'message': '下載成功', 'file_id': new_file.id})
        # except:
        #     return json.dumps({'error': 1, 'message': '下載失敗'})

    @http.route('/other_equipment/down_wrong_file', type='http', auth="user", methods=['GET'])
    def down_wrong_file(self, **kwargs):
        '''
        返回错误excel内容
        :param kwargs: file_id
        :return:
        '''
        file_id = int(kwargs['file_id'])
        filetype = kwargs['type']
        wb = request.env['maintenance_plan.trans.excel'].browse(file_id)
        response = request.make_response(wb.file)
        if filetype == '錯誤':
            response.headers["Content-Disposition"] = "attachment; filename={}". \
                format((wb.name + '错误.xlsx').encode().decode('latin-1'))
        if filetype == '下載':
            response.headers["Content-Disposition"] = "attachment; filename={}". \
                format((wb.name + '.xlsx').encode().decode('latin-1'))
        return response
