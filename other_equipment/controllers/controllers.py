# -*- coding: utf-8 -*-
import json
import os
import random
import time
import xlrd
import xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.writer.excel import save_virtual_workbook

from odoo import http
from odoo.http import request
from xlutils3.copy import copy as xl_copy
from datetime import datetime as dt

ROW_1_LIST = ['EQUIPMENT No.', 'TEAM NO.', 'RESULT \nREFERENCE', 'EQUIPMENT',
              'BRAND', 'MODEL', 'SERIAL_NO', 'MANUAL REF. NO.',
              'EQUIPMENT OWNER', 'LOCATION OF EQUIPMENT', 'FREQ. OF CAL.', 'CALIBRATION BODY',
              'CALIBRATION REQUIPEMNETS  (SEE NOTE)', 'LAST MAINTENANCE DATE', 'MAINTENANCE DUE DATE', 'STATUS',
              'REMARK']
APP_DIR = os.path.dirname(os.path.dirname(__file__))


class OtherEquipment(http.Controller):
    @staticmethod
    def excel_validate(request, new_sheet, num, cols_list, style, has_date_col):
        '''
        驗證excel的一行的錯誤性
        :param sheet: sheet活動表
        :param num: 行數，1開始
        :param cols_list: 需要驗證的列數列表
        :param style: sheet顏色style
        :param has_date_col: 是否驗證第17、18列的時間格式
        :return:
        '''
        row_error = False
        for col in cols_list:

            if new_sheet.cell(num, col).value is None:
                new_sheet.cell(num, col).fill = style
                row_error = True
            elif (col == 14 or col == 15) and has_date_col is True:
                try:
                    print()
                    dt.strptime(new_sheet.cell(num, col).value, '%Y/%m/%d %H:%M')
                except Exception as e:
                    new_sheet.cell(num, col).fill = style
                    row_error = True
                    print(new_sheet.cell(num, col).value)
                    print('这个时间格式不对')
            else:
                if col == 1:
                    # 設備編號
                    equipment_num = new_sheet.cell(num, col).value
                    # 組號
                    if new_sheet.cell(num, 2).value == '':
                        row_error = True
                    else:
                        departments = new_sheet.cell(num, 2).value
                        # 檢查當前組是否有這個設備編號了
                        if request.env['other_equipment.other_equipment'].sudo().search_count(
                                [('equipment_num', '=', equipment_num),
                                 ('departments.department_order', '=', departments)]) > 0:
                            row_error = True
                        else:
                            new_sheet.cell(num, col).fill = style
                # 驗證這個組是否存在
                if col == 2:
                    departments = new_sheet.cell(num, 2).value
                    if not request.env['user.department'].sudo().search([
                             ('department_order', '=', departments)]):
                        row_error = True
                        print(new_sheet.cell(num, col).value)
                        print('没得这个组的')
                    else:
                        new_sheet.cell(num, col).fill = style
        return row_error

    @staticmethod
    def set_excel_style_color(colour_map_key):
        '''
        設置單元格背景色
        :param colour_map_key: 顏色對應鍵值對的key
        :return:
        '''
        style = xlwt.XFStyle()
        pattern = xlwt.Pattern()
        pattern.pattern = 1
        pattern.pattern_fore_colour = xlwt.Style.colour_map[colour_map_key]
        style.pattern = pattern
        return style

    @http.route('/other_equipment/put_in_excel', type='http', csrf=False, auth='user')
    def put_in_excel(self, **kwargs):
        '''
        工器具管理導入excel按鈕
        :param kwargs: excel的file信息
        :return:
        '''
        file = kwargs['file']
        filename = kwargs['file'].filename
        workbook = load_workbook(file)
        sheet = workbook.active
        # 设置单元格背景色紅色
        red_style = PatternFill(fill_type='solid', fgColor="FF3030")
        # 设置单元格背景色綠色
        green_style = PatternFill(fill_type='solid', fgColor="458B74")
        error = False
        n_row = 0
        for row in sheet.rows:
            n_row += 1
            if n_row == 1:
                if [i.value for i in row] != ROW_1_LIST:
                    return json.dumps({'message': '表格不符', 'error': True})
            else:
                # 這裡是檢測日期格式是否正確
                row_error = self.excel_validate(request, sheet, n_row, [14, 15], red_style, has_date_col=True)
                # 這裡檢查機器編號是否合格
                row_error2 = self.excel_validate(request, sheet, n_row,
                                                 [1, 2, 4, 10, 11, 16], red_style, has_date_col=False)
                if row_error or row_error2:
                    error = True
                else:
                    # 設備編號
                    equipment_num = sheet.cell(n_row, 1).value
                    # 所屬群組
                    departments = request.env['user.department'].search([
                             ('department_order', '=', sheet.cell(n_row, 2).value)]).id
                    # 結果參考
                    result_reference = sheet.cell(n_row, 3).value
                    # 設備名稱
                    equipment_name = sheet.cell(n_row, 4).value
                    # 品牌
                    brand = sheet.cell(n_row, 5).value
                    # 型號
                    model = sheet.cell(n_row, 6).value
                    # 序列號
                    serial_no = sheet.cell(n_row, 7).value
                    # 參考手冊編號
                    manual_ref_no = sheet.cell(n_row, 8).value
                    # 設備擁有者
                    equipment_owner = sheet.cell(n_row, 9).value
                    # 設備位置
                    location_of_equipment = sheet.cell(n_row, 10).value
                    # 檢驗週期
                    freq_of_cal = sheet.cell(n_row, 11).value
                    # 檢驗主體
                    calibration_body = sheet.cell(n_row, 12).value
                    # 檢驗要求
                    calibration_requipemnets = sheet.cell(n_row, 13).value
                    # 最後維護日期
                    last_maintenance_date = sheet.cell(n_row, 14).value
                    # 應用到期時間
                    maintenance_due_data = sheet.cell(n_row, 15).value
                    # 狀態
                    status = sheet.cell(n_row, 16).value
                    # 備註
                    remark = sheet.cell(n_row, 17).value
                    request.env['other_equipment.other_equipment'].create({
                        'departments': departments, 'equipment_num': equipment_num,
                        'result_reference': result_reference, 'equipment_name': equipment_name,
                        'brand': brand, 'model': model,'serial_no': serial_no, 'manual_ref_no': manual_ref_no,
                        'equipment_owner': equipment_owner, 'location_of_equipment': location_of_equipment,
                        'freq_of_cal': freq_of_cal, 'calibration_body': calibration_body,
                        'calibration_requipemnets': calibration_requipemnets,
                        'last_maintenance_date': last_maintenance_date, 'status': status,
                        'maintenance_due_data': maintenance_due_data, 'remark': remark,
                    })
                    error = False
                if row_error is True:
                    error = True
        if error is True:
            new_file = request.env['other_equipment.trans.excel'].create({
                'name': filename.split('.')[0],
                'file': save_virtual_workbook(workbook)
            })
            return json.dumps({'error': error, 'message': '文件有部分錯誤信息，請修改后再次傳入', 'file_id': new_file.id})
        else:
            return json.dumps({'error': error, 'message': '上傳成功'})


    @http.route('/other_equipment/get_in_excel', type='http', csrf=False, auth='user')
    def get_in_excel(self, **kwargs):
        '''
        工器具管理導出excel按鈕
        :param kwargs: 需要導出的domian條件
        :return:
        '''
        domains = kwargs['domains']
        arr = []
        try:
            if domains:
                x = domains.split(',')
                num = int(len(x)/3 + 1)
                for i in range(1, num):
                    tuples = (x[3*i-3], x[3*i-2], x[3*i-1])
                    arr.append(tuples)
                other_equipments = request.env['other_equipment.other_equipment'].search(arr)
            else:
                other_equipments = request.env['other_equipment.other_equipment'].search([])
            wb = Workbook()
            # 激活 worksheet
            ws = wb.active
            ws.append(ROW_1_LIST)
            len_other_equipments = len(other_equipments)
            for i in range(1, len_other_equipments+1):
                my_records = [
                    other_equipments[i-1].equipment_num if other_equipments[i-1].equipment_num else '',
                    other_equipments[i - 1].departments.department_order if other_equipments[
                        i - 1].departments.department_order else '',
                    other_equipments[i - 1].result_reference if other_equipments[i - 1].result_reference else '',
                    other_equipments[i - 1].equipment_name if other_equipments[i - 1].equipment_name else '',
                    other_equipments[i - 1].brand if other_equipments[i - 1].brand else '',
                    other_equipments[i - 1].model if other_equipments[i - 1].model else '',
                    other_equipments[i - 1].serial_no if other_equipments[i - 1].serial_no else '',
                    other_equipments[i - 1].manual_ref_no if other_equipments[i - 1].manual_ref_no else '',
                    other_equipments[i - 1].equipment_owner if other_equipments[i - 1].equipment_owner else '',
                    other_equipments[i - 1].location_of_equipment if other_equipments[
                        i - 1].location_of_equipment else '',
                    other_equipments[i - 1].freq_of_cal if other_equipments[i - 1].freq_of_cal else '',
                    other_equipments[i - 1].calibration_body if other_equipments[i - 1].calibration_body else '',
                    other_equipments[i - 1].calibration_requipemnets if other_equipments[
                        i - 1].calibration_requipemnets else '',
                    other_equipments[i - 1].last_maintenance_date if other_equipments[
                        i - 1].last_maintenance_date else '',
                    other_equipments[i - 1].maintenance_due_data if other_equipments[i - 1].maintenance_due_data else '',
                    other_equipments[i - 1].status if other_equipments[i - 1].status else '',
                    other_equipments[i - 1].remark if other_equipments[i - 1].remark else '',
                ]
                ws.append(my_records)
            wb.save("sample.xlsx")
            with open('sample.xlsx', 'rb') as f:
                data = f.read()
                new_file = request.env['other_equipment.trans.excel'].create({
                    'name': '工器具導出數據',
                    'file': data
                })
                os.remove('sample.xlsx')
            return json.dumps({'error': 0, 'message': '下載成功', 'file_id': new_file.id})
        except:
            return json.dumps({'error': 1, 'message': '下載失敗'})

    @http.route('/other_equipment/down_wrong_file', type='http', auth="user", methods=['GET'])
    def down_wrong_file(self, **kwargs):
        '''
        返回错误excel内容
        :param kwargs: file_id
        :return:
        '''
        file_id = int(kwargs['file_id'])
        filetype = kwargs['type']
        wb = request.env['other_equipment.trans.excel'].browse(file_id)
        response = request.make_response(wb.file)
        if filetype == '錯誤':
            response.headers["Content-Disposition"] = "attachment; filename={}". \
                format((wb.name + '错误.xlsx').encode().decode('latin-1'))
        if filetype == '下載':
            response.headers["Content-Disposition"] = "attachment; filename={}". \
                format((wb.name + '.xlsx').encode().decode('latin-1'))
        return response